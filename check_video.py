#-*- coding: utf-8 -*-
# check availability of video urls
# availability is composed of three settings; embeddable, regionRestriction, syndicated
# the first two settings are checked in this program
# the last setting called syndicated is not checked because it is accessed only by search api that can have only 1 video id
# and it is rare case that a content is not syndicated and but embeddable and region allowed
# region restriction need country code and it is hard coded as KR for now.

import os
import sys
from Tkinter import *
import Tkinter, Tkconstants, tkFileDialog
import pdb
from openpyxl import load_workbook, Workbook
reload(sys)
sys.setdefaultencoding('utf-8')

# need to install some packages
# pip install --upgrade google-api-python-client
# pip install --upgrade oauth2client

from apiclient.discovery import build
from apiclient.errors import HttpError
from oauth2client.tools import argparser

# Set DEVELOPER_KEY to the API key value from the APIs & auth > Registered apps tab of https://cloud.google.com/console
# Please ensure that you have enabled the YouTube Data API for your project.

with open("youtube_api.key") as fp:
  DEVELOPER_KEY = fp.readline().strip()
YOUTUBE_API_SERVICE_NAME = "youtube"
YOUTUBE_API_VERSION = "v3"

# accept single video id only
def check_syndicated(video_id):
  youtube = build(YOUTUBE_API_SERVICE_NAME, YOUTUBE_API_VERSION, developerKey=DEVELOPER_KEY)
  response = youtube.search().list(q=video_id, type="video", part="id", videoSyndicated="true").execute()
  if int(response['pageInfo']['totalResults']) == 0:
    return False
  else:
    return True

# accept comma-separated multiple video id
def check_region(video_id, country_code):
  video_ids = video_id.split(",")
  youtube = build(YOUTUBE_API_SERVICE_NAME, YOUTUBE_API_VERSION, developerKey=DEVELOPER_KEY)
  response = youtube.videos().list(id=video_id, part="contentDetails", maxResults=50).execute()  # regionRestriction
  items = response['items']
  if len(video_ids) <= 1:
    if len(items) > 0:
      result = True
      if 'regionRestriction' in items[0]['contentDetails']:
        regions = items[0]['contentDetails']['regionRestriction']
        if 'allowed' in regions:
          if regions['allowed'].count(country_code) == 0:
            result = False
        if 'blocked' in regions:
          if regions['blocked'].count(country_code) > 0:
            result = False
      return result
    else:
      return -1
  else:
    result_array = []
    j = 0
    for video_id in video_ids:
      if video_id == items[j]['id']:
        result = True
        if 'regionRestriction' in items[j]['contentDetails']:
          regions = items[j]['contentDetails']['regionRestriction']
          if 'allowed' in regions:
            if regions['allowed'].count(country_code) == 0:
              result = False
          if 'blocked' in regions:
            if regions['blocked'].count(country_code) > 0:
              result = False
        result_array.append(result)
        j += 1
        if j == len(items): j -= 1
      else:
        result_array.append(-1)
    return result_array

# accept comma-separated multiple video id
def check_embeddable(video_id):
  video_ids = video_id.split(",")
  youtube = build(YOUTUBE_API_SERVICE_NAME, YOUTUBE_API_VERSION, developerKey=DEVELOPER_KEY)
  response = youtube.videos().list(id=video_id, part="status", maxResults=50).execute()  # embeddable
  items = response['items']
  if len(video_ids) <= 1:
    if len(items) > 0:
      return items[0]['status']['embeddable']
    else:
      return -1
  else:
    result_array = []
    j = 0
    for video_id in video_ids:
      if video_id == items[j]['id']:
        result_array.append(items[j]['status']['embeddable'])
        j += 1
        if j == len(items): j -= 1
      else:
        result_array.append(-1)
    return result_array

# make api call and write result
def run_checking(infos, is_excel, outp):
  video_ids = ""
  for info in infos:
    videos = info[3]
    for video in videos:
      video_id = video.split("=")[-1]
      video_ids += video_id + ","
  video_ids = video_ids[:-1]

  embeddables = check_embeddable(video_ids)
  restricteds = check_region(video_ids, "KR")

  j = 0
  for info in infos:
    serial = info[0]
    series = info[1]
    book = info[2]
    videos = info[3]
    offset = 1
    for video in videos:
      if embeddables[j] == False or embeddables[j] == -1 or restricteds[j] == False:
        if is_excel:
          outp.append([book, str(offset), video, str(embeddables[j]), str(restricteds[j])])
        else:
          outp.write(book + "\t" + str(offset) + "\t" + video + "\t" + str(embeddables[j]) + "\t" + str(restricteds[j]) + "\n")
      offset += 1
      j += 1

def read_excel(xlsx):
  lines = []
  wb = load_workbook(xlsx)
  ws = wb.active
  for row in ws.rows:
    serial = str(row[0].value)
    series = row[1].value
    book = row[2].value
    url1 = row[3].value
    url2 = row[4].value
    if url2 == None: url2 = ""
    url3 = row[5].value
    if url3 == None: url3 = ""
    if book != "":
      lines.append(serial + "\t" + series + "\t" + book + "\t" + url1 + "\t" + url2 + "\t" + url3)
  return lines

if __name__ == "__main__":
  if len(sys.argv) < 2:
    root = Tk()
    root.filename = tkFileDialog.askopenfilename(title = "Select Video List Excel File", filetypes = (("excel files", "*.xlsx"), ("all files", "*.*")))
    infile = root.filename
    if len(infile) == 0:
      print "excel file not selected"
      sys.exit()
  else:
    infile = sys.argv[1]
  is_excel = infile.endswith("xlsx")
  if is_excel:
    lines = read_excel(infile)
    outfile = infile[:-5] + "_check.xlsx"
  else:
    vlf = open(infile, 'r')
    lines = vlf.readlines()
    vlf.close()
    outfile = infile[:-4] + "_check.tsv"

  num_books = 0
  num_videos = 0
  if is_excel:
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(['book', 'index', 'video', 'embeddable', 'region'])
  else:
    outp = open(outfile, 'w')
    outp.write("book\tindex\tvideo\tembeddable\tregion\n")

  # call youtube api every 10 books
  infos = []
  for line in lines:
    line = line.decode('utf-8').strip()
    if line == "": continue

    tokens = line.split("\t")
    serial = tokens[0]
    series = tokens[1]
    book = tokens[2]
    videos = []
    for i in range(3,len(tokens)):
      videos.append(tokens[i])
      num_videos += 1
    num_books += 1
    infos.append([serial, series, book, videos])
    if num_books % 10 == 0:
      print "now checking", num_books, "/", len(lines)
      if is_excel:
        run_checking(infos, is_excel, ws2)
      else:
        run_checking(infos, is_excel, outp)
      infos = []
  if len(infos) > 0:
    if is_excel:
      run_checking(infos, is_excel, ws2)
    else:
      run_checking(infos, is_excel, outp)
  
  if is_excel:
    wb2.save(outfile)
  else:
    outp.close()
